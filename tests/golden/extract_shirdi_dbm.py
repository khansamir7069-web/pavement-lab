"""Extract the input data and expected outputs from the source .xlsm
into a JSON fixture that tests/test_excel_parity.py uses.

Run this once to produce shirdi_dbm.json (already committed).
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl

SRC = Path(__file__).resolve().parents[3] / "source_files" / "Marshal Mix final 2 (1).xlsm"
OUT = Path(__file__).with_name("shirdi_dbm.json")


def main() -> None:
    wb = openpyxl.load_workbook(SRC, data_only=True)

    g = wb["Gradation"]
    sp = wb["Sp.Gr."]
    gmb = wb["Gmb"]
    gmm = wb["Gmm"]
    sf = wb["Stability and Flow"]
    charts = wb["Charts"]
    report = wb["Report"]
    title = wb["Title"]
    mat = wb["Material  Cal"]   # NB: sheet name has 2 spaces

    sieve_sizes = tuple(g.cell(r, 1).value for r in range(8, 16))
    aggs = {
        "25mm": tuple(g.cell(r, 2).value for r in range(8, 16)),
        "20mm": tuple(g.cell(r, 3).value for r in range(8, 16)),
        "6mm":  tuple(g.cell(r, 4).value for r in range(8, 16)),
        "SD":   tuple(g.cell(r, 5).value for r in range(8, 16)),
        "Cement": tuple(g.cell(r, 6).value for r in range(8, 16)),
    }
    blend = {
        "25mm": g.cell(7, 7).value,
        "20mm": g.cell(7, 8).value,
        "6mm":  g.cell(7, 9).value,
        "SD":   g.cell(7, 10).value,
        "Cement": g.cell(7, 11).value,
    }
    spec_low = tuple(g.cell(r, 13).value for r in range(8, 16))
    spec_up  = tuple(g.cell(r, 14).value for r in range(8, 16))

    # Expected combined gradation
    expected_combined = tuple(g.cell(r, 12).value for r in range(8, 16))

    # Specific gravity inputs — 25mm: rows 8-11 cols C-F (4 reps)
    def coarse_block(row_a, row_b, row_c, row_d, cols):
        return {
            "A": tuple(sp.cell(row_a, c).value for c in cols),
            "B": tuple(sp.cell(row_b, c).value for c in cols),
            "C": tuple(sp.cell(row_c, c).value for c in cols),
            "D": tuple(sp.cell(row_d, c).value for c in cols),
        }

    # Only first column is filled in the source file for 25mm and 20mm
    sg_25 = coarse_block(8, 9, 10, 11, [3])
    sg_20 = coarse_block(25, 26, 27, 28, [3])

    # 6mm pycnometer: rows 8-11 cols K-N
    def fine_block(rows, cols):
        return {
            "W1": tuple(sp.cell(rows[0], c).value for c in cols),
            "W2": tuple(sp.cell(rows[1], c).value for c in cols),
            "W3": tuple(sp.cell(rows[2], c).value for c in cols),
            "W4": tuple(sp.cell(rows[3], c).value for c in cols),
        }

    sg_6mm = fine_block([8, 9, 10, 11], [11])
    sg_sd = fine_block([19, 20, 21, 22], [11])

    # Bitumen rows 8-11 cols S-V
    sg_bit = {
        "A": tuple(sp.cell(8, c).value for c in [19]),
        "B": tuple(sp.cell(9, c).value for c in [19]),
        "C": tuple(sp.cell(10, c).value for c in [19]),
        "D": tuple(sp.cell(11, c).value for c in [19]),
    }

    # Gmb: per-Pb groups of 3 specimens. Pb at column B header for each group start.
    gmb_groups = []
    for start_row, pb in [(6, 3.5), (9, 4.0), (12, 4.5), (15, 5.0), (18, 5.5)]:
        specs = []
        for r in range(start_row, start_row + 3):
            specs.append({
                "A": gmb.cell(r, 3).value,
                "D": gmb.cell(r, 4).value,
                "B": gmb.cell(r, 5).value,
            })
        gmb_groups.append({"pb_pct": pb, "specimens": specs})

    # Gmm reference samples at Pb=4.5%
    gmm_ref = []
    for r in [5, 6]:
        gmm_ref.append({
            "A": gmm.cell(r, 3).value,
            "B": gmm.cell(r, 4).value,
            "D": gmm.cell(r, 6).value,
            "E": gmm.cell(r, 7).value,
        })

    # Stability and flow: 5 groups of 3 samples. Per-group avg formulas:
    #   Pb=3.5 (rows 5-7): AVERAGE(N5:N7) all 3 ; flow AVERAGE(P5:P7) all 3
    #   Pb=4.0 (rows 9-11): AVERAGE(N10:N11) sample 2 & 3 ; flow AVERAGE(P9:P11)
    #   Pb=4.5 (rows 13-15): (N14+N15)/2 sample 2 & 3 ; flow AVERAGE(P13:P15)
    #   Pb=5.0 (rows 17-19): (N17+N19)/2 sample 1 & 3 ; flow AVERAGE(P17:P19)
    #   Pb=5.5 (rows 21-23): AVERAGE(N21:N23) all 3   ; flow (P21+P23)/2 sample 1 & 3
    stab_include = {
        (5, 3.5): [True, True, True],
        (9, 4.0): [False, True, True],
        (13, 4.5): [False, True, True],
        (17, 5.0): [True, False, True],
        (21, 5.5): [True, True, True],
    }
    flow_include = {
        (5, 3.5): [True, True, True],
        (9, 4.0): [True, True, True],
        (13, 4.5): [True, True, True],
        (17, 5.0): [True, True, True],
        (21, 5.5): [True, False, True],
    }
    sf_groups = []
    for start_row, pb in [(5, 3.5), (9, 4.0), (13, 4.5), (17, 5.0), (21, 5.5)]:
        specs = []
        for idx, r in enumerate(range(start_row, start_row + 3)):
            specs.append({
                "h1": sf.cell(r, 3).value,
                "h2": sf.cell(r, 4).value,
                "h3": sf.cell(r, 5).value,
                "dia": sf.cell(r, 8).value,
                "corr": sf.cell(r, 12).value,
                "stab": sf.cell(r, 13).value,
                "flow": sf.cell(r, 16).value,
                "n_cached": sf.cell(r, 14).value,    # cached Corrected Stability
                "include_stab": stab_include[(start_row, pb)][idx],
                "include_flow": flow_include[(start_row, pb)][idx],
            })
        sf_groups.append({"pb_pct": pb, "specimens": specs})

    # Expected outputs (from Charts sheet)
    expected_summary = []
    for r in range(2, 7):
        expected_summary.append({
            "pb": charts.cell(r, 1).value,
            "agg": charts.cell(r, 2).value,
            "gmm": charts.cell(r, 3).value,
            "gmb": charts.cell(r, 4).value,
            "air_voids": charts.cell(r, 5).value,
            "vma": charts.cell(r, 6).value,
            "vfb": charts.cell(r, 7).value,
            "stability": charts.cell(r, 8).value,
            "flow": charts.cell(r, 9).value,
            "mq": charts.cell(r, 10).value,
        })
    gsb_expected = charts.cell(11, 4).value

    # Expected Sp.Gr. derived values
    sp_expected = {
        "Gb_avg": sp.cell(13, 18).value,            # R13: avg SG bitumen
        "25mm_bulk_od_avg": sp.cell(17, 3).value,   # C17
        "20mm_bulk_od_avg": sp.cell(34, 3).value,   # C34
        "6mm_sg_avg": sp.cell(15, 10).value,        # J15
        "SD_sg_avg": sp.cell(26, 10).value,         # J26
        "Gsb": sp.cell(37, 10).value,               # J37
    }

    # Gmm derived
    gmm_expected = {
        "Gmm_avg_at_ref": gmm.cell(7, 8).value,      # H7
        "Gse": gmm.cell(8, 8).value,                  # H8
        "per_pb": [
            {"pb": gmm.cell(r, 2).value, "gmm": gmm.cell(r, 4).value}
            for r in range(14, 19)
        ],
    }

    # Material Calculation (sheet "Material  Cal")
    material_calc = {
        "inputs": {
            "standard_bitumen_pct": mat.cell(6, 4).value,       # D6 = 4.5
            "standard_aggregate_weight_g": mat.cell(9, 5).value, # E9 = 1200
            "target_bitumen_pct": mat.cell(6, 9).value,         # I6 = 4
        },
        "standard_expected": {
            "aggregate_pct": mat.cell(9, 4).value,              # D9
            "bitumen_weight_g": mat.cell(10, 5).value,          # E10
            "total_mix_pct": mat.cell(11, 4).value,             # D11
            "total_mix_weight_g": mat.cell(11, 5).value,        # E11
            "aggregate_weight_restated_g": mat.cell(13, 4).value,  # D13
            "bitumen_weight_restated_g": mat.cell(14, 4).value,    # D14
            "total_bituminous_mix_g": mat.cell(15, 4).value,       # D15
        },
        "target_expected": {
            "aggregate_pct": mat.cell(9, 9).value,              # I9
            "aggregate_weight_g": mat.cell(9, 10).value,        # J9
            "bitumen_weight_g": mat.cell(10, 10).value,         # J10
            "total_mix_pct": mat.cell(11, 9).value,             # I11
            "total_mix_weight_g": mat.cell(11, 10).value,       # J11
        },
        "dry_material_standard_expected": [
            {"name": "25mm",   "pct": mat.cell(19, 4).value, "weight_g": mat.cell(19, 5).value},
            {"name": "20mm",   "pct": mat.cell(20, 4).value, "weight_g": mat.cell(20, 5).value},
            {"name": "6mm",    "pct": mat.cell(21, 4).value, "weight_g": mat.cell(21, 5).value},
            {"name": "SD",     "pct": mat.cell(22, 4).value, "weight_g": mat.cell(22, 5).value},
            {"name": "Cement", "pct": mat.cell(23, 4).value, "weight_g": mat.cell(23, 5).value},
        ],
        "dry_material_standard_total_expected": mat.cell(24, 5).value,
        "dry_material_target_expected": [
            {"name": "25mm",   "pct": mat.cell(19, 9).value, "weight_g": mat.cell(19, 10).value},
            {"name": "20mm",   "pct": mat.cell(20, 9).value, "weight_g": mat.cell(20, 10).value},
            {"name": "6mm",    "pct": mat.cell(21, 9).value, "weight_g": mat.cell(21, 10).value},
            {"name": "SD",     "pct": mat.cell(22, 9).value, "weight_g": mat.cell(22, 10).value},
            {"name": "Cement", "pct": mat.cell(23, 9).value, "weight_g": mat.cell(23, 10).value},
        ],
        "dry_material_target_total_expected": mat.cell(24, 10).value,
    }

    fixture = {
        "title": {
            "A2": title.cell(2, 1).value,
            "mix_type": title.cell(4, 4).value,
        },
        "gradation": {
            "sieve_sizes_mm": sieve_sizes,
            "pass_pct": aggs,
            "blend_ratios": blend,
            "spec_lower": spec_low,
            "spec_upper": spec_up,
            "expected_combined": expected_combined,
        },
        "sp_gr": {
            "coarse_25": sg_25,
            "coarse_20": sg_20,
            "fine_6": sg_6mm,
            "fine_sd": sg_sd,
            "bitumen": sg_bit,
            "expected": sp_expected,
        },
        "gmb": {"groups": gmb_groups},
        "gmm": {"reference_pb": 4.5, "samples_ref": gmm_ref, "expected": gmm_expected},
        "stability_flow": {"groups": sf_groups},
        "summary_expected": expected_summary,
        "gsb_expected": gsb_expected,
        "obc_expected": report.cell(4, 2).value,    # B4 = 0.04254 (4.254%)
        "material_calc": material_calc,
    }

    OUT.write_text(json.dumps(fixture, indent=2, default=str), encoding="utf-8")
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
